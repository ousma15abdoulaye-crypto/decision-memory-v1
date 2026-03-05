#!/usr/bin/env python3
"""
Probe format réel fichiers IMC.
RÈGLE-30 : on ne suppose pas le format · on le lit.
Stratégie : local-first · LlamaCloud = dernier recours.

Usage :
    $env:DATABASE_URL        = "<Railway>"
    $env:LLAMA_CLOUD_API_KEY = "<clé>"  # uniquement si PDF scan
    python scripts/probe_imc_format.py
"""

import json
import os
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data" / "imports" / "imc"


def probe_csv(file: Path) -> None:
    print("FORMAT : CSV")
    lines = file.read_text(encoding="utf-8", errors="replace").splitlines()
    print(f"Lignes totales : {len(lines)}")
    for line in lines[:10]:
        print(f"  {repr(line)}")


def probe_excel(file: Path) -> None:
    print("FORMAT : Excel")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        print(f"Feuilles : {wb.sheetnames}")
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            print(f"  {row}")
    except ImportError:
        print("❌ openpyxl manquant · pip install openpyxl")


def probe_pdf_native(file: Path) -> bool:
    """
    Tente extraction texte natif PDF.
    Retourne True si texte récupéré · False si scan (image).
    """
    try:
        import pdfminer.high_level as pdf

        text = pdf.extract_text(str(file))
        if text and len(text.strip()) > 100:
            print("FORMAT : PDF texte natif ✓ (pas besoin LlamaCloud)")
            print(f"500 premiers caractères :\n{text[:500]}")
            return True
        else:
            print("FORMAT : PDF scan (texte natif vide)")
            return False
    except ImportError:
        print("pdfminer.six non installé · pip install pdfminer.six")
        return False


def probe_pdf_llama(file: Path) -> None:
    """Fallback LlamaCloud si PDF scan."""
    api_key = os.environ.get("LLAMA_CLOUD_API_KEY")
    if not api_key:
        print("❌ PDF scan détecté mais LLAMA_CLOUD_API_KEY absente")
        print("   Export requis pour continuer le probe")
        return

    cache_dir = REPO_ROOT / ".llama_cache"
    cache_dir.mkdir(exist_ok=True)
    cache_key = __import__("hashlib").sha256(file.read_bytes()).hexdigest()
    cache_file = cache_dir / f"{cache_key}_imc_probe.json"

    if cache_file.exists():
        print("FORMAT : PDF via LlamaCloud (CACHE HIT)")
        data = json.loads(cache_file.read_text())
        print(data["text"][:500])
        return

    print("FORMAT : PDF via LlamaCloud (CACHE MISS · 1 appel API)")
    from llama_parse import LlamaParse

    parser = LlamaParse(api_key=api_key, result_type="markdown", language="fr")
    docs = parser.load_data(str(file))
    if docs:
        text = docs[0].text
        cache_file.write_text(json.dumps({"text": text}))
        print(f"Pages    : {len(docs)}")
        print(f"500 chars :\n{text[:500]}")
    else:
        print("VIDE")


def main() -> None:
    if not DATA_DIR.exists():
        print(f"❌ Dossier {DATA_DIR} introuvable")
        print("   Créer data/imports/imc/ et y placer les fichiers IMC")
        sys.exit(1)

    files = sorted(
        [
            f
            for f in DATA_DIR.rglob("*")
            if f.suffix.lower() in {".pdf", ".xlsx", ".xls", ".csv"}
        ]
    )

    if not files:
        print(f"❌ Aucun fichier dans {DATA_DIR}")
        print("   Copier les PDF/CSV/Excel IMC dans ce dossier")
        sys.exit(1)

    years = sorted(
        {
            int(m.group(1))
            for f in files
            for m in [re.search(r"(201[89]|202[0-6])", f.name)]
            if m
        }
    )

    print("=" * 60)
    print("PROBE IMC FORMAT")
    print(f"Fichiers trouvés : {len(files)}")
    print(f"Années détectées : {years}")
    print(f"Fichier test     : {files[0].name}")
    print("=" * 60)

    f = files[0]
    ext = f.suffix.lower()

    if ext == ".csv":
        probe_csv(f)
    elif ext in {".xlsx", ".xls"}:
        probe_excel(f)
    elif ext == ".pdf":
        native_ok = probe_pdf_native(f)
        if not native_ok:
            probe_pdf_llama(f)

    print("\n" + "=" * 60)
    print("PROBE TERMINÉ · Poster le résultat au CTO · attendre GO")
    print("=" * 60)


if __name__ == "__main__":
    main()
