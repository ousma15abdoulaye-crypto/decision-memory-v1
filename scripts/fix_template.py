#!/usr/bin/env python3
"""
Script de normalisation du template CBA canonique
Objectif: Supprimer toute ambiguïté Excel sans modifier la logique métier

Usage:
    python scripts/fix_template.py <chemin_template.xlsx>
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.workbook import Workbook


def list_sheetnames(wb: Workbook, label: str) -> None:
    """Affiche les noms d'onglets avec un label"""
    print(f"\n{label}:")
    for idx, name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. '{name}'")


def normalize_template(template_path: Path) -> None:
    """
    Normalise un template CBA:
    1. Liste les sheetnames AVANT
    2. Renomme les onglets avec espaces/suffixes
    3. Supprime ou masque les onglets debug
    4. Sauvegarde le template
    5. Reliste les sheetnames APRÈS
    """
    if not template_path.exists():
        print(f"❌ ERREUR: Template introuvable: {template_path}")
        sys.exit(1)

    print(f"📋 Normalisation du template: {template_path.name}")
    print(f"   Chemin: {template_path}")

    # Charger le workbook
    wb = openpyxl.load_workbook(template_path)

    # 1. Lister AVANT
    list_sheetnames(wb, "ONGLETS AVANT NORMALISATION")

    # 2. Normalisation des noms d'onglets
    # Mapping des corrections à appliquer
    sheet_renames = {
        "Commercial Evaluation ": "Commercial Evaluation",  # Espace en fin
        "Commercial Evaluation (2)": "Commercial Evaluation",  # Suffixe (2)
        "Essential Evaluation ": "Essential Evaluation",
        "Capability Evaluation ": "Capability Evaluation",
        "Sustainability Evaluation ": "Sustainability Evaluation",
        "Summary ": "Summary",
    }

    # Détecter et renommer
    modifications = []
    for old_name, new_name in sheet_renames.items():
        if old_name in wb.sheetnames:
            ws = wb[old_name]
            ws.title = new_name
            modifications.append(f"  ✓ Renommé: '{old_name}' → '{new_name}'")

    if modifications:
        print("\n🔧 MODIFICATIONS APPLIQUÉES:")
        for mod in modifications:
            print(mod)
    else:
        print("\n✓ Aucune normalisation de nom nécessaire")

    # 3. Masquer/supprimer onglets debug
    debug_sheets = ["DMS_SUMMARY", "DEBUG", "TEMP", "SCRATCH", "NOTES"]
    hidden_sheets = []

    for sheet_name in list(
        wb.sheetnames
    ):  # Liste copie pour éviter modification pendant itération
        # Vérifier si c'est un onglet debug (exact ou pattern)
        is_debug = False
        for debug_pattern in debug_sheets:
            if debug_pattern in sheet_name.upper():
                is_debug = True
                break

        if is_debug:
            ws = wb[sheet_name]
            # On masque plutôt que supprimer pour préserver les données
            ws.sheet_state = "hidden"
            hidden_sheets.append(sheet_name)

    if hidden_sheets:
        print("\n🔒 ONGLETS DEBUG MASQUÉS:")
        for name in hidden_sheets:
            print(f"  ✓ Masqué: '{name}'")
    else:
        print("\n✓ Aucun onglet debug détecté")

    # 4. Sauvegarder
    backup_path = template_path.with_suffix(".backup.xlsx")
    print(f"\n💾 Sauvegarde de l'original: {backup_path.name}")
    wb.save(backup_path)

    print(f"💾 Sauvegarde du template normalisé: {template_path.name}")
    wb.save(template_path)

    # 5. Lister APRÈS
    wb_after = openpyxl.load_workbook(template_path)
    list_sheetnames(wb_after, "ONGLETS APRÈS NORMALISATION")

    print("\n✅ Normalisation terminée avec succès")


def main():
    if len(sys.argv) != 2:
        print("Usage: python scripts/fix_template.py <chemin_template.xlsx>")
        print("\nExemple:")
        print(
            "  python scripts/fix_template.py src/templates/DMS-CBA-CANONICAL-V1.0.xlsx"
        )
        sys.exit(1)

    template_path = Path(sys.argv[1])
    normalize_template(template_path)


if __name__ == "__main__":
    main()
