"""Unit tests for styling module."""
from unittest.mock import MagicMock

from openpyxl.styles import PatternFill

from src.mapping.styling import CONFIDENCE_COLORS, apply_confidence_styling


class TestConfidenceColors:
    """Tests for CONFIDENCE_COLORS dictionary."""

    def test_high_confidence_no_color(self):
        """High confidence should have no color (None)."""
        assert CONFIDENCE_COLORS["high"] is None

    def test_medium_confidence_has_color(self):
        """Medium confidence should have a yellow-ish color."""
        assert CONFIDENCE_COLORS["medium"] == "FFF3CD"

    def test_low_confidence_has_color(self):
        """Low confidence should have an orange-ish color."""
        assert CONFIDENCE_COLORS["low"] == "FFE5CC"

    def test_human_only_has_color(self):
        """Human-only should have same color as low."""
        assert CONFIDENCE_COLORS["human_only"] == "FFE5CC"

    def test_error_has_color(self):
        """Error should have a red-ish color."""
        assert CONFIDENCE_COLORS["error"] == "F8D7DA"


class TestApplyConfidenceStyling:
    """Tests for apply_confidence_styling function."""

    def test_high_confidence_no_fill_applied(self):
        """High confidence should not modify cell fill."""
        cell = MagicMock()
        original_fill = cell.fill

        apply_confidence_styling(cell, "high")

        # Fill should not be changed (no assignment)
        assert cell.fill == original_fill

    def test_medium_confidence_applies_fill(self):
        """Medium confidence should apply yellow fill."""
        cell = MagicMock()

        apply_confidence_styling(cell, "medium")

        # Check that fill was set
        assert cell.fill is not None
        call_args = cell.fill
        assert isinstance(call_args, PatternFill)
        assert call_args.start_color.rgb == "00FFF3CD"  # openpyxl prepends 00

    def test_low_confidence_applies_fill(self):
        """Low confidence should apply orange fill."""
        cell = MagicMock()

        apply_confidence_styling(cell, "low")

        assert isinstance(cell.fill, PatternFill)
        assert cell.fill.start_color.rgb == "00FFE5CC"

    def test_human_only_applies_fill(self):
        """Human-only should apply same fill as low."""
        cell = MagicMock()

        apply_confidence_styling(cell, "human_only")

        assert isinstance(cell.fill, PatternFill)
        assert cell.fill.start_color.rgb == "00FFE5CC"

    def test_error_applies_fill(self):
        """Error should apply red fill."""
        cell = MagicMock()

        apply_confidence_styling(cell, "error")

        assert isinstance(cell.fill, PatternFill)
        assert cell.fill.start_color.rgb == "00F8D7DA"

    def test_unknown_level_no_fill_applied(self):
        """Unknown confidence level should not modify cell fill."""
        cell = MagicMock()
        original_fill = cell.fill

        apply_confidence_styling(cell, "unknown_level")

        # Fill should not be changed
        assert cell.fill == original_fill

    def test_empty_string_level_no_fill_applied(self):
        """Empty string confidence level should not modify cell fill."""
        cell = MagicMock()
        original_fill = cell.fill

        apply_confidence_styling(cell, "")

        assert cell.fill == original_fill

    def test_fill_type_is_solid(self):
        """Applied fills should use solid fill type."""
        cell = MagicMock()

        apply_confidence_styling(cell, "medium")

        assert cell.fill.fill_type == "solid"

    def test_start_and_end_color_match(self):
        """Start and end colors should match for solid fill."""
        cell = MagicMock()

        apply_confidence_styling(cell, "error")

        assert cell.fill.start_color.rgb == cell.fill.end_color.rgb
