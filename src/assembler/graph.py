"""LangGraph state machine Pass -1 — ZIP → bundles fournisseurs.

Pipeline stateful :
  extract (ZIP) → classify (M12) → bundle (groupement) → hitl_check → finalize

Référence : ADR-V420-002-LANGGRAPH.md
Performance : ZIP 15 fichiers SCI Mali → 4 bundles < 30s.
HITL : bundle incomplet → interrupt() → reprise sans re-OCR.
"""

from __future__ import annotations

import logging
import os
import tempfile
import zipfile
from pathlib import Path
from typing import TypedDict

from src.observability.pipeline_v5_metrics import observe_pass1_hitl_bypass

logger = logging.getLogger(__name__)

try:
    from langgraph.graph import END, StateGraph
    from langgraph.types import interrupt

    _LANGGRAPH_AVAILABLE = True
except ImportError:
    _LANGGRAPH_AVAILABLE = False
    logger.warning("[ASSEMBLER] langgraph non installé — graph indisponible.")


class PassMinusOneState(TypedDict):
    """État LangGraph du Pass -1."""

    workspace_id: str
    tenant_id: str
    zip_path: str
    extract_dir: str
    raw_documents: list[dict]
    bundles_draft: list[dict]
    hitl_required: bool
    hitl_resolved: bool
    finalized: bool
    bundle_ids: list[str]
    error: str | None


async def extract_node(state: PassMinusOneState) -> PassMinusOneState:
    """Nœud 1 : Extraire les fichiers du ZIP + OCR."""
    from src.assembler.ocr_azure import ocr_with_azure
    from src.assembler.ocr_mistral import ocr_native_pdf, ocr_with_mistral
    from src.assembler.pdf_detector import FileType, detect_file_type
    from src.assembler.zip_validator import validate_zip

    zip_path = state["zip_path"]
    validation = validate_zip(zip_path)
    if not validation.is_valid:
        return {**state, "error": validation.error, "finalized": True}

    extract_dir = tempfile.mkdtemp(prefix="dms_pass_minus1_")
    raw_documents = []

    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_dir)

    for filename in validation.filenames:
        file_path = Path(extract_dir) / filename
        if not file_path.exists():
            continue

        file_type = detect_file_type(file_path)

        if file_type == FileType.NATIVE_PDF:
            ocr_result = await ocr_native_pdf(file_path)
        elif file_type in {FileType.SCAN, FileType.IMAGE}:
            ocr_result = await ocr_with_mistral(file_path)
            if ocr_result.get("error"):
                ocr_result = await ocr_with_azure(file_path)
        elif file_type == FileType.WORD:
            ocr_result = _extract_word(file_path)
        elif file_type == FileType.EXCEL:
            ocr_result = _extract_excel(file_path)
        else:
            ocr_result = {"raw_text": "", "confidence": 0.0, "ocr_engine": "none"}

        raw_documents.append(
            {
                "filename": filename,
                "storage_path": str(file_path),
                "file_type": file_type.value,
                "ocr_result": ocr_result,
                "raw_text": ocr_result.get("raw_text", ""),
            }
        )

    return {**state, "extract_dir": extract_dir, "raw_documents": raw_documents}


async def classify_node(state: PassMinusOneState) -> PassMinusOneState:
    """Nœud 2 : Classification M12 de chaque document."""
    classified = []
    for doc in state.get("raw_documents", []):
        classified.append(
            {
                **doc,
                "doc_type": _classify_document_type(
                    doc.get("raw_text", ""), doc.get("filename", "")
                ),
                "doc_role": "primary",
                "m12_doc_kind": _classify_document_type(
                    doc.get("raw_text", ""), doc.get("filename", "")
                ),
                "m12_confidence": 0.8,
            }
        )

    return {**state, "raw_documents": classified}


async def bundle_node(state: PassMinusOneState) -> PassMinusOneState:
    """Nœud 3 : Groupement des documents en bundles par fournisseur."""
    bundles: dict[str, list[dict]] = {}

    for doc in state.get("raw_documents", []):
        vendor = resolve_bundle_vendor_key(
            doc.get("raw_text", ""), doc.get("filename", "")
        )
        if vendor not in bundles:
            bundles[vendor] = []
        bundles[vendor].append(doc)

    bundles_draft = []
    hitl_required = False

    for vendor_name, docs in bundles.items():
        completeness, missing = _check_completeness(docs)
        needs_hitl = completeness < 0.6
        if needs_hitl:
            hitl_required = True

        bundles_draft.append(
            {
                "vendor_name_raw": vendor_name,
                "documents": docs,
                "completeness_score": completeness,
                "missing_documents": missing,
                "hitl_required": needs_hitl,
            }
        )

    return {**state, "bundles_draft": bundles_draft, "hitl_required": hitl_required}


async def hitl_check_node(state: PassMinusOneState) -> PassMinusOneState:
    """Nœud 4 : Interruption HITL si bundle incomplet."""
    if state.get("hitl_required") and not state.get("hitl_resolved"):
        # E2E / CI headless : sans reprise humaine, ``interrupt()`` empêche ``finalize``
        # → 0 bundle en base. Activer explicitement ``DMS_PASS1_HEADLESS=1`` uniquement
        # pour scripts locaux ou jobs non interactifs (jamais par défaut en prod).
        if os.environ.get("DMS_PASS1_HEADLESS", "").strip().lower() in (
            "1",
            "true",
            "yes",
        ):
            observe_pass1_hitl_bypass(workspace_id=str(state.get("workspace_id", "")))
            return {**state, "hitl_resolved": True}
        incomplete = [
            b for b in state.get("bundles_draft", []) if b.get("hitl_required")
        ]
        resolved = interrupt(
            {
                "workspace_id": state["workspace_id"],
                "incomplete_bundles": [b["vendor_name_raw"] for b in incomplete],
                "message": "Bundles incomplets — validation manuelle requise.",
            }
        )
        return {**state, "hitl_resolved": True, **resolved}
    return state


async def finalize_node(state: PassMinusOneState) -> PassMinusOneState:
    """Nœud 5 : Écriture finale en base — supplier_bundles + bundle_documents."""
    from src.assembler.bundle_writer import write_bundle

    bundle_ids = []
    for idx, bundle_draft in enumerate(state.get("bundles_draft", [])):
        bid = write_bundle(
            workspace_id=state["workspace_id"],
            tenant_id=state["tenant_id"],
            vendor_name_raw=bundle_draft["vendor_name_raw"],
            bundle_index=idx,
            documents=bundle_draft["documents"],
            completeness_score=bundle_draft["completeness_score"],
            missing_documents=bundle_draft["missing_documents"],
            hitl_required=bundle_draft.get("hitl_required", False),
        )
        bundle_ids.append(bid)

    return {**state, "finalized": True, "bundle_ids": bundle_ids}


def _route_hitl(state: PassMinusOneState) -> str:
    if state.get("hitl_required") and not state.get("hitl_resolved"):
        return "hitl"
    return "ok"


def classify_document_type_for_pass_minus_one(raw_text: str, filename: str) -> str:
    """Return the existing Pass-1 M12 document kind for raw text and filename."""
    text_lower = (raw_text + " " + filename).lower()
    if any(k in text_lower for k in ["offre", "prix", "montant", "total"]):
        return "offer_combined"
    if "nif" in text_lower or "identification fiscale" in text_lower:
        return "nif"
    if "rccm" in text_lower or "registre" in text_lower:
        return "rccm"
    if "rib" in text_lower or "relevé d'identité" in text_lower:
        return "rib"
    return "other"


def _classify_document_type(raw_text: str, filename: str) -> str:
    return classify_document_type_for_pass_minus_one(raw_text, filename)


def resolve_bundle_vendor_key(raw_text: str, zip_entry_filename: str) -> str:
    """Clé de regroupement Pass -1 : dossier racine du ZIP si chemin multi-segments.

    Convention pilote : ``FournisseurA/doc.docx`` → un bundle **FournisseurA** (stable,
    aligné « N offres = N dossiers »). Fichiers à la racine du ZIP : heuristique
    texte / nom de fichier inchangée (``_extract_vendor_name``).
    """
    name = (zip_entry_filename or "").replace("\\", "/").strip()
    parts = [p for p in name.split("/") if p and p not in (".", "..")]
    if len(parts) >= 2:
        return parts[0][:80]
    return _extract_vendor_name(raw_text, zip_entry_filename)


def _extract_vendor_name(raw_text: str, filename: str) -> str:
    for line in (raw_text or "").split("\n")[:20]:
        line = line.strip()
        if len(line) > 5 and any(
            k in line.upper()
            for k in ["SARL", "SA ", "ETS", "SAS", "CABINET", "BV ", "COOP"]
        ):
            return line[:80]
    parts = Path(filename).stem.split("_")
    if parts:
        return parts[0][:80]
    return "FOURNISSEUR_INCONNU"


def _check_completeness(docs: list[dict]) -> tuple[float, list[str]]:
    required = {"offer_combined", "nif", "rccm"}
    present = {d.get("doc_type", "other") for d in docs}
    missing = sorted(required - present)
    score = len(required - set(missing)) / len(required)
    return score, missing


def _extract_word(file_path: Path) -> dict:
    try:
        from docx import Document  # type: ignore[import-untyped]

        doc = Document(str(file_path))
        text = "\n".join(p.text for p in doc.paragraphs)
        return {"raw_text": text, "confidence": 1.0, "ocr_engine": "none"}
    except Exception as exc:
        return {
            "raw_text": "",
            "confidence": 0.0,
            "ocr_engine": "none",
            "error": str(exc),
        }


def _extract_excel(file_path: Path) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(str(file_path), read_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                lines.append(" ".join(str(c) for c in row if c is not None))
        return {"raw_text": "\n".join(lines), "confidence": 1.0, "ocr_engine": "none"}
    except Exception as exc:
        return {
            "raw_text": "",
            "confidence": 0.0,
            "ocr_engine": "none",
            "error": str(exc),
        }


def build_pass_minus_one_graph():
    """Construit et compile le graphe LangGraph du Pass -1.

    Returns:
        CompiledGraph prêt à l'utilisation, ou None si langgraph non installé.
    """
    if not _LANGGRAPH_AVAILABLE:
        logger.error("[ASSEMBLER] langgraph requis pour build_pass_minus_one_graph().")
        return None

    graph = StateGraph(PassMinusOneState)

    graph.add_node("extract", extract_node)
    graph.add_node("classify", classify_node)
    graph.add_node("bundle", bundle_node)
    graph.add_node("hitl_check", hitl_check_node)
    graph.add_node("finalize", finalize_node)

    graph.set_entry_point("extract")
    graph.add_edge("extract", "classify")
    graph.add_edge("classify", "bundle")
    graph.add_conditional_edges(
        "bundle",
        _route_hitl,
        {"hitl": "hitl_check", "ok": "finalize"},
    )
    graph.add_edge("hitl_check", "finalize")
    graph.add_edge("finalize", END)

    return graph.compile()
