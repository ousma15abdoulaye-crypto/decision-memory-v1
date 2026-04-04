"""CBA export and review services for Couche A."""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.db.workspace_context import get_workspace_id_for_case

from ..models import (
    deserialize_json,
    ensure_schema,
    generate_id,
    get_connection,
    serialize_json,
)

ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


def _fill_manual_review(cell) -> None:
    cell.fill = ORANGE_FILL


def _build_workbook(payload: dict[str, Any]) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Offre", payload.get("supplier_name", "Inconnu")])
    summary.append(["Montant", payload.get("amount")])
    summary.append(["Statut", payload.get("status")])

    for sheet_name in ["Essentiels", "Capacité", "Durabilité", "Commercial"]:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Critère", "Valeur"])
        for key, value in payload.get(sheet_name.lower(), {}).items():
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=key)
            cell = ws.cell(row=row, column=2, value=value)
            if payload.get("manual_review"):
                _fill_manual_review(cell)
    return wb


async def export_cba(offer_id: str, actor: str | None = None) -> dict[str, Any]:
    """Generate and store a CBA export for an offer."""

    def _process() -> dict[str, Any]:
        ensure_schema()
        with get_connection() as conn:
            conn.execute(
                "SELECT * FROM offers WHERE id = %(offer_id)s",
                {"offer_id": offer_id},
            )
            offer = conn.fetchone()
            if not offer:
                raise ValueError("Offre introuvable.")

            conn.execute(
                """
                SELECT * FROM extractions
                WHERE offer_id = %(offer_id)s
                ORDER BY created_at DESC NULLS LAST, id DESC
                LIMIT 1
                """,
                {"offer_id": offer_id},
            )
            extraction = conn.fetchone()

            conn.execute(
                """
                SELECT * FROM analyses
                WHERE offer_id = %(offer_id)s
                ORDER BY id DESC
                LIMIT 1
                """,
                {"offer_id": offer_id},
            )
            analysis = conn.fetchone()

            missing_fields = []
            if extraction:
                missing_fields = deserialize_json(extraction.get("missing_json"))
                if not isinstance(missing_fields, list):
                    missing_fields = []
            manual_review = not extraction or bool(missing_fields)

            conn.execute(
                """
                SELECT COUNT(*) AS cnt FROM documents
                WHERE offer_id = %(offer_id)s AND document_type = 'CBA_EXPORT'
                """,
                {"offer_id": offer_id},
            )
            row = conn.fetchone()
            existing_exports = row["cnt"] if row else 0
            version = int(existing_exports) + 1

            payload = {
                "supplier_name": offer["supplier_name"],
                "amount": offer.get("amount"),
                "status": analysis["status"] if analysis else "EN_ATTENTE",
                "essentiels": {"fournisseur": offer["supplier_name"]},
                "capacité": {},
                "durabilité": {},
                "commercial": {"montant": offer.get("amount")},
                "manual_review": manual_review,
            }
            wb = _build_workbook(payload)

            export_dir = Path("data") / "couche_a" / "exports" / offer_id
            export_dir.mkdir(parents=True, exist_ok=True)
            filename = f"cba_{offer_id}_v{version}.xlsx"
            export_path = export_dir / filename
            wb.save(export_path)

            doc_id = generate_id()
            metadata = {"version": version, "manual_review": manual_review}
            storage_path = str(export_path)
            _ws_id = get_workspace_id_for_case(conn, offer["case_id"])
            _ws_clause = ", workspace_id" if _ws_id else ""
            _ws_val = ", %(workspace_id)s" if _ws_id else ""
            conn.execute(
                f"""
                INSERT INTO documents (
                    id, case_id, lot_id, offer_id, document_type, filename,
                    storage_path, path, mime_type, metadata_json, uploaded_at{_ws_clause}
                )
                VALUES (
                    %(id)s, %(case_id)s, %(lot_id)s, %(offer_id)s, 'CBA_EXPORT', %(filename)s,
                    %(storage_path)s, %(path)s, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    %(metadata_json)s, NOW()::TEXT{_ws_val}
                )
                """,
                {
                    "id": doc_id,
                    "case_id": offer["case_id"],
                    "lot_id": offer.get("lot_id"),
                    "offer_id": offer_id,
                    "filename": filename,
                    "storage_path": storage_path,
                    "path": storage_path,
                    "metadata_json": serialize_json(metadata),
                    "workspace_id": _ws_id,
                },
            )
            conn.execute(
                """
                INSERT INTO audits (
                    id, case_id, entity_type, entity_id, action, actor, details_json, created_at
                )
                VALUES (
                    %(id)s, %(case_id)s, 'cba', %(entity_id)s, 'EXPORT', %(actor)s, %(details_json)s, NOW()::TEXT
                )
                """,
                {
                    "id": generate_id(),
                    "case_id": offer["case_id"],
                    "entity_id": doc_id,
                    "actor": actor,
                    "details_json": serialize_json(metadata),
                },
            )

        return {
            "document_id": doc_id,
            "offer_id": offer_id,
            "version": version,
            "path": str(export_path),
        }

    return await asyncio.to_thread(_process)


async def upload_cba_review(
    offer_id: str, upload: UploadFile, reviewer: str | None = None
) -> dict[str, Any]:
    """Store a completed CBA upload for committee review."""
    if not upload.filename:
        raise ValueError("Le fichier fourni est invalide.")

    file_bytes = await upload.read()
    filename = Path(upload.filename).name
    review_dir = Path("data") / "couche_a" / "cba_reviews" / offer_id
    review_dir.mkdir(parents=True, exist_ok=True)
    review_path = review_dir / filename

    def _persist() -> dict[str, Any]:
        ensure_schema()
        review_path.write_bytes(file_bytes)
        with get_connection() as conn:
            conn.execute(
                "SELECT * FROM offers WHERE id = %(offer_id)s",
                {"offer_id": offer_id},
            )
            offer = conn.fetchone()
            if not offer:
                raise ValueError("Offre introuvable.")
            doc_id = generate_id()
            _ws_id_rv = get_workspace_id_for_case(conn, offer["case_id"])
            _rv_ws_clause = ", workspace_id" if _ws_id_rv else ""
            _rv_ws_val = ", %(workspace_id)s" if _ws_id_rv else ""
            conn.execute(
                f"""
                INSERT INTO documents (
                    id, case_id, lot_id, offer_id, document_type, filename,
                    storage_path, path, mime_type, metadata_json, uploaded_at{_rv_ws_clause}
                )
                VALUES (
                    %(id)s, %(case_id)s, %(lot_id)s, %(offer_id)s, 'CBA_REVIEW', %(filename)s,
                    %(storage_path)s, %(path)s, %(mime_type)s, %(metadata_json)s, NOW()::TEXT{_rv_ws_val}
                )
                """,
                {
                    "id": doc_id,
                    "case_id": offer["case_id"],
                    "lot_id": offer.get("lot_id"),
                    "offer_id": offer_id,
                    "filename": filename,
                    "storage_path": str(review_path),
                    "path": str(review_path),
                    "mime_type": upload.content_type or "application/octet-stream",
                    "metadata_json": serialize_json({"reviewer": reviewer}),
                    "workspace_id": _ws_id_rv,
                },
            )
            conn.execute(
                """
                INSERT INTO audits (
                    id, case_id, entity_type, entity_id, action, actor, details_json, created_at
                )
                VALUES (
                    %(id)s, %(case_id)s, 'cba', %(entity_id)s, 'REVIEW_UPLOAD', %(actor)s, %(details_json)s, NOW()::TEXT
                )
                """,
                {
                    "id": generate_id(),
                    "case_id": offer["case_id"],
                    "entity_id": doc_id,
                    "actor": reviewer,
                    "details_json": serialize_json({"filename": filename}),
                },
            )
        return {"document_id": doc_id, "path": str(review_path)}

    return await asyncio.to_thread(_persist)
