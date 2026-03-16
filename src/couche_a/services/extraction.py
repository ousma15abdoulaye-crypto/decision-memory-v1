"""Document extraction service for Couche A."""

from __future__ import annotations

import asyncio
import re
from pathlib import Path
from typing import Any

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from ..models import (
    ensure_schema,
    generate_id,
    get_connection,
    serialize_json,
)


def _extract_text_from_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages[:3]]
    return "\n".join(pages)


def _extract_text_from_docx(path: Path) -> str:
    doc = DocxDocument(str(path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs)


def _extract_text_from_xlsx(path: Path) -> str:
    workbook = load_workbook(str(path), read_only=True)
    text_lines: list[str] = []
    for sheet in workbook.worksheets[:1]:
        for row in sheet.iter_rows(max_row=20, max_col=10):
            for cell in row:
                if cell.value:
                    text_lines.append(str(cell.value))
    workbook.close()
    return "\n".join(text_lines)


def extract_text(path: Path) -> str:
    """Extract text from supported document types."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_text_from_pdf(path)
    if suffix in {".docx", ".doc"}:
        return _extract_text_from_docx(path)
    if suffix in {".xlsx", ".xls"}:
        return _extract_text_from_xlsx(path)
    return path.read_text(errors="ignore")


def _normalize_amount(raw_amount: str) -> float:
    cleaned = raw_amount.replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def extract_fields(text: str) -> tuple[dict[str, Any], list[str]]:
    """Extract structured fields from raw text."""
    missing_fields: list[str] = []
    supplier_match = re.search(
        r"(fournisseur|soumissionnaire)\s*[:\-]\s*([^\n]+)", text, re.IGNORECASE
    )
    supplier_name = supplier_match.group(2).strip() if supplier_match else None

    date_match = re.search(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})", text)
    depot_date = date_match.group(1) if date_match else None

    amount_match = re.search(
        r"(\d{1,3}(?:[\s,.]\d{3})*(?:[,.]\d{2})?)\s?(fcfa|eur|usd)?",
        text,
        re.IGNORECASE,
    )
    amount = _normalize_amount(amount_match.group(1)) if amount_match else None

    zone_match = re.search(r"zone\s*[:\-]\s*([^\n]+)", text, re.IGNORECASE)
    zone = zone_match.group(1).strip() if zone_match else None

    attachments = [
        line.strip() for line in text.splitlines() if "annexe" in line.lower()
    ]

    fields = {
        "fournisseur": supplier_name,
        "date_depot": depot_date,
        "montant": amount,
        "documents_joints": attachments,
        "zone": zone,
    }

    for key, value in fields.items():
        if value in (None, "", []):
            missing_fields.append(key)

    return fields, missing_fields


async def extract_and_store(
    document_id: str, llm_enabled: bool = False
) -> dict[str, Any]:
    """Extract data from a document and store the results."""

    def _process() -> dict[str, Any]:
        ensure_schema()
        with get_connection() as conn:
            conn.execute(
                "SELECT * FROM documents WHERE id = %(document_id)s",
                {"document_id": document_id},
            )
            doc_row = conn.fetchone()
            if not doc_row:
                raise ValueError("Document introuvable.")

            storage_path = doc_row.get("storage_path") or doc_row.get("path")
            if not storage_path:
                raise ValueError("Document sans chemin de stockage.")
            file_path = Path(storage_path)
            text = extract_text(file_path)
            extracted, missing = extract_fields(text)

            extraction_id = generate_id()
            offer_id = doc_row.get("offer_id")
            conn.execute(
                """
                INSERT INTO extractions (
                    id, case_id, document_id, offer_id,
                    extracted_json, missing_json, used_llm, data_json, created_at
                )
                VALUES (
                    %(id)s, %(case_id)s, %(document_id)s, %(offer_id)s,
                    %(extracted_json)s, %(missing_json)s, %(used_llm)s, %(data_json)s, NOW()::TEXT
                )
                """,
                {
                    "id": extraction_id,
                    "case_id": doc_row["case_id"],
                    "document_id": document_id,
                    "offer_id": offer_id,
                    "extracted_json": serialize_json(extracted),
                    "missing_json": serialize_json(missing),
                    "used_llm": llm_enabled,
                    "data_json": serialize_json(extracted),
                },
            )

            if offer_id and (
                extracted.get("fournisseur") is not None
                or extracted.get("montant") is not None
            ):
                supplier = extracted.get("fournisseur")
                montant = extracted.get("montant")
                if supplier is not None and montant is not None:
                    conn.execute(
                        "UPDATE offers SET supplier_name = %(supplier_name)s, amount = %(amount)s WHERE id = %(offer_id)s",
                        {
                            "supplier_name": supplier,
                            "amount": montant,
                            "offer_id": offer_id,
                        },
                    )
                elif supplier is not None:
                    conn.execute(
                        "UPDATE offers SET supplier_name = %(supplier_name)s WHERE id = %(offer_id)s",
                        {"supplier_name": supplier, "offer_id": offer_id},
                    )
                else:
                    conn.execute(
                        "UPDATE offers SET amount = %(amount)s WHERE id = %(offer_id)s",
                        {"amount": montant, "offer_id": offer_id},
                    )

            conn.execute(
                """
                INSERT INTO audits (
                    id, case_id, entity_type, entity_id, action, actor, details_json, created_at
                )
                VALUES (
                    %(id)s, %(case_id)s, 'extraction', %(entity_id)s, 'EXTRACTION', NULL, %(details_json)s, NOW()::TEXT
                )
                """,
                {
                    "id": generate_id(),
                    "case_id": doc_row["case_id"],
                    "entity_id": extraction_id,
                    "details_json": serialize_json(
                        {"missing_fields": missing, "llm_enabled": llm_enabled}
                    ),
                },
            )

        return {
            "extraction_id": extraction_id,
            "document_id": document_id,
            "extracted": extracted,
            "missing_fields": missing,
            "llm_enabled": llm_enabled,
        }

    return await asyncio.to_thread(_process)
