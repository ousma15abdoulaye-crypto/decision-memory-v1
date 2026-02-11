"""
CBA Excel Generator – Save the Children Format (5 onglets)
Constitution V2.1 § 4.1 – Manuel SCI SC-PR-02 § 5.3
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional

# === CHARTE SAVE THE CHILDREN ===
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Arial")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def generate_cba_excel(case_data: Dict, output_dir: Path) -> Path:
    """
    Génère le fichier CBA Excel avec 5 onglets selon format SCI.

    Args:
        case_data: Dict avec clés:
            - case_reference: str
            - market_title: str
            - dao_ref: str
            - opening_date: datetime
            - lot: str (optional)
            - suppliers: List[Dict] avec metadata offres
            - technical_criteria: List[Dict] (nom, ponderation)
            - financial_offers: List[Dict] (supplier, montant_ht, tva, delai)
    Returns:
        Path du fichier généré
    """
    wb = Workbook()

    # Onglet 1: Informations Générales
    _build_info_sheet(wb, case_data)

    # Onglet 2: Registre Dépôt
    _build_registre_sheet(wb, case_data)

    # Onglet 3: Analyse Technique
    _build_technique_sheet(wb, case_data)

    # Onglet 4: Analyse Financière
    _build_financiere_sheet(wb, case_data)

    # Onglet 5: Synthèse
    _build_synthese_sheet(wb, case_data)

    # Supprimer sheet par défaut
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    filename = f"CBA_{case_data['case_reference']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    return filepath


def _build_info_sheet(wb: Workbook, data: Dict):
    """Onglet 1: Informations Générales"""
    ws = wb.create_sheet("Info Générales", 0)

    # Titre fusionné
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = "COMPARATIVE BID ANALYSIS (CBA)"
    title.font = Font(size=16, bold=True, color="366092", name="Arial")
    title.alignment = Alignment(horizontal='center')

    # Section infos marché
    infos = [
        ("Titre du marché:", data.get('market_title', '')),
        ("Référence DAO/RFQ:", data.get('dao_ref', '')),
        ("Lot:", data.get('lot', 'Lot unique')),
        ("Date ouverture:", data['opening_date'].strftime('%d/%m/%Y') if data.get('opening_date') else ''),
        ("Nombre soumissions:", len(data.get('suppliers', []))),
    ]

    row = 3
    for label, value in infos:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:F{row}')
        row += 1

    # Section Membres Comité (3 lignes vides)
    row += 2
    ws[f'A{row}'] = "Membres comité d'évaluation:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for i in range(1, 4):
        ws[f'A{row}'] = f"{i}."
        ws[f'B{row}'] = "Nom: _______________________"
        ws[f'D{row}'] = "Fonction: _________________"
        ws[f'F{row}'] = "Signature: _______________"
        row += 1

    # Bordures
    for r in range(1, row):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER


def _build_registre_sheet(wb: Workbook, data: Dict):
    """Onglet 2: Registre Dépôt"""
    ws = wb.create_sheet("Registre Dépôt")

    # Headers
    headers = ["N°", "Soumissionnaire", "Date Dépôt", "Heure Dépôt",
               "Mode Dépôt", "Docs Reçus", "Observation"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Données
    suppliers = data.get('suppliers', [])
    # Tri chronologique
    suppliers.sort(key=lambda x: x.get('timestamp', datetime.now()))

    for idx, sup in enumerate(suppliers, start=1):
        row = idx + 1
        ws.cell(row, 1, idx)
        ws.cell(row, 2, sup.get('supplier_name', ''))
        ts = sup.get('timestamp', datetime.now())
        ws.cell(row, 3, ts.strftime('%d/%m/%Y'))
        ws.cell(row, 4, ts.strftime('%H:%M:%S'))
        ws.cell(row, 5, sup.get('mode_depot', 'Email'))
        ws.cell(row, 6, sup.get('docs_count', 0))
        ws.cell(row, 7, '')

        for col in range(1, 8):
            ws.cell(row, col).border = THIN_BORDER

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30


def _normalize_criteria_key(name: str) -> str:
    """Normalise le nom du critère pour la recherche (sans accents)."""
    import unicodedata
    nfd = unicodedata.normalize('NFD', name.lower().replace(' ', '_'))
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')


def _build_technique_sheet(wb: Workbook, data: Dict):
    """Onglet 3: Analyse Technique (dynamique selon DAO)"""
    ws = wb.create_sheet("Analyse Technique")

    criteria = data.get('technical_criteria', [])
    if not criteria:
        criteria = [{"name": "Capacité technique", "weight": 50},
                   {"name": "Durabilité", "weight": 10}]

    # Headers
    # Ligne 1: Noms critères
    ws.cell(1, 1, "Soumissionnaire")
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT

    for idx, crit in enumerate(criteria, start=2):
        cell = ws.cell(1, idx, crit['name'])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        # Ligne 2: Pondération
        ws.cell(2, idx, f"/{crit.get('weight', 0)} pts")
        ws.cell(2, idx).fill = SUBHEADER_FILL
        ws.cell(2, idx).font = Font(italic=True, size=9)

    # Total Technique
    total_col = len(criteria) + 2
    ws.cell(1, total_col, "Total Technique")
    ws.cell(1, total_col).fill = HEADER_FILL
    ws.cell(1, total_col).font = HEADER_FONT

    # Champs humains
    ws.cell(1, total_col + 1, "Visite Fournisseur")
    ws.cell(1, total_col + 1).fill = WARNING_FILL
    ws.cell(1, total_col + 1).font = Font(bold=True)

    ws.cell(1, total_col + 2, "Évaluation Échantillon")
    ws.cell(1, total_col + 2).fill = WARNING_FILL
    ws.cell(1, total_col + 2).font = Font(bold=True)

    # Lignes fournisseurs
    suppliers = data.get('suppliers', [])
    for row_idx, sup in enumerate(suppliers, start=3):
        ws.cell(row_idx, 1, sup.get('supplier_name', ''))

        # Extraire données techniques (mock)
        tech_data = sup.get('extracted_data', {}).get('technique', {})
        for col_idx, crit in enumerate(criteria, start=2):
            key = _normalize_criteria_key(crit['name'])
            val = tech_data.get(key) or tech_data.get(crit['name'].lower().replace(' ', '_'), '')
            ws.cell(row_idx, col_idx, val)
            ws.cell(row_idx, col_idx).border = THIN_BORDER

        # Formule Total
        last_crit_letter = get_column_letter(len(criteria) + 1)
        ws.cell(row_idx, total_col, f"=SUM(B{row_idx}:{last_crit_letter}{row_idx})")
        ws.cell(row_idx, total_col).border = THIN_BORDER

        # Champs humains vides
        ws.cell(row_idx, total_col + 1, '')
        ws.cell(row_idx, total_col + 2, '')


def _build_financiere_sheet(wb: Workbook, data: Dict):
    """Onglet 4: Analyse Financière"""
    ws = wb.create_sheet("Analyse Financière")

    headers = ["Soumissionnaire", "Montant HT (FCFA)", "TVA (%)",
               "Montant TTC (FCFA)", "Délai Livraison (jours)", "Note Financière"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    suppliers = data.get('suppliers', [])
    for row_idx, sup in enumerate(suppliers, start=2):
        fin_data = sup.get('extracted_data', {}).get('financiere', {})
        ws.cell(row_idx, 1, sup.get('supplier_name', ''))
        ws.cell(row_idx, 2, fin_data.get('montant_ht', 0))
        ws.cell(row_idx, 3, fin_data.get('tva', 0))
        # TTC = HT * (1 + TVA/100)
        ws.cell(row_idx, 4, f"=B{row_idx}*(1+C{row_idx}/100)")
        ws.cell(row_idx, 5, fin_data.get('delai_jours', 0))
        ws.cell(row_idx, 6, '')  # Note (humain)

        for col in range(1, 7):
            ws.cell(row_idx, col).border = THIN_BORDER


def _build_synthese_sheet(wb: Workbook, data: Dict):
    """Onglet 5: Synthèse – Classement final"""
    ws = wb.create_sheet("Synthèse")

    headers = ["Classement", "Soumissionnaire", "Note Technique (/100)",
               "Note Financière (/100)", "Note Finale", "Observations Comité"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    suppliers = data.get('suppliers', [])
    criteria = data.get('technical_criteria', [])
    total_tech_col = len(criteria) + 2
    tech_col_letter = get_column_letter(total_tech_col)

    for row_idx, sup in enumerate(suppliers, start=2):
        technique_row = row_idx + 1  # Données technique commencent à row 3
        ws.cell(row_idx, 1, row_idx - 1)  # Classement provisoire
        ws.cell(row_idx, 2, sup.get('supplier_name', ''))
        ws.cell(row_idx, 3, f"='Analyse Technique'!{tech_col_letter}{technique_row}")
        ws.cell(row_idx, 4, f"='Analyse Financière'!F{row_idx}")
        ws.cell(row_idx, 5, f"=C{row_idx}*0.7+D{row_idx}*0.3")  # Pondération 70/30
        ws.cell(row_idx, 6, '')

        for col in range(1, 7):
            ws.cell(row_idx, col).border = THIN_BORDER

    # Formule de classement
    last_row = len(suppliers) + 1
    for row in range(2, last_row + 1):
        ws.cell(row, 1, f"=RANK(E{row}, $E$2:$E${last_row}, 0)")
