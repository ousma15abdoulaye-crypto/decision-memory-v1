from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .column_calculator import commercial_cols, supplier_col_letter
from .styling import apply_confidence_styling


def populate_summary_supplier(
    ws: Worksheet, spec: dict, slot: int, supplier_name: str
) -> None:
    s = spec["sheets"]["Summary"]
    col = supplier_col_letter(s["supplier_start_col_index"], slot, width_per_slot=1)

    # Label + Nom
    ws[f"{col}{s['supplier_label_row']}"] = f"Soumissionnaire {slot:02d}"
    ws[f"{col}{s['supplier_name_row']}"] = supplier_name

    # Liens scores (rows 6-9) + total row 10
    links = s["score_links"]
    sr = s["score_rows"]

    def target_col_letter(link_key: str) -> str:
        lk = links[link_key]
        start = lk["target_start_col_index"]
        width = lk["col_offset"]
        # essentials/capability/sustainability => width=1; commercial => width=2 (points to TOTAL col)
        idx = start + (slot - 1) * width
        from openpyxl.utils import get_column_letter

        return get_column_letter(idx)

    # Essentials
    e_col = target_col_letter("essentials")
    ws[f"{col}{sr['essentials']}"] = links["essentials"]["formula_pattern"].format(
        col=e_col, row=links["essentials"]["target_row"]
    )
    # Capability
    c_col = target_col_letter("capability")
    ws[f"{col}{sr['capability']}"] = links["capability"]["formula_pattern"].format(
        col=c_col, row=links["capability"]["target_row"]
    )
    # Sustainability
    su_col = target_col_letter("sustainability")
    ws[f"{col}{sr['sustainability']}"] = links["sustainability"][
        "formula_pattern"
    ].format(col=su_col, row=links["sustainability"]["target_row"])
    # Commercial
    co_col = target_col_letter("commercial")
    ws[f"{col}{sr['commercial']}"] = links["commercial"]["formula_pattern"].format(
        col=co_col, row=links["commercial"]["target_row"]
    )

    # Total
    essentials_cell = f"{col}{sr['essentials']}"
    cap_cell = f"{col}{sr['capability']}"
    sust_cell = f"{col}{sr['sustainability']}"
    comm_cell = f"{col}{sr['commercial']}"
    ws[f"{col}{sr['total']}"] = s["total_score_formula"].format(
        essentials_cell=essentials_cell,
        cap_cell=cap_cell,
        sust_cell=sust_cell,
        comm_cell=comm_cell,
    )


def populate_essential_supplier(
    ws: Worksheet, spec: dict, slot: int, conformity_data: dict[str, Any]
) -> None:
    s = spec["sheets"]["Essential Evaluation"]
    col = supplier_col_letter(s["supplier_start_col_index"], slot, width_per_slot=1)

    # Label
    ws[f"{col}{s['supplier_label_row']}"] = f"Soumissionnaire {slot:02d}"

    # Nom (référence Summary)
    summary_col = supplier_col_letter(
        spec["sheets"]["Summary"]["supplier_start_col_index"], slot, 1
    )
    ws[f"{col}{s['supplier_name_row']}"] = s["name_formula_pattern"].format(
        summary_col=summary_col
    )

    # Critères : si data absente => "Pass" low confidence (à valider)
    _ = s["criteria_start_row"]
    for k in range(s["criteria_start_row"], s["criteria_end_row"] + 1):
        # On mappe par ordre d'apparition des clés si fourni
        pass

    keys = list(conformity_data.keys())
    for idx, r in enumerate(range(s["criteria_start_row"], s["criteria_end_row"] + 1)):
        if idx < len(keys):
            ok = bool(conformity_data[keys[idx]])
            ws[f"{col}{r}"] = "Pass" if ok else "Fail"
        else:
            ws[f"{col}{r}"] = "Pass"
            apply_confidence_styling(ws[f"{col}{r}"], "medium")

    # Score final
    ws[f"{col}{s['score_row']}"] = s["score_formula"].format(col=col)


def populate_capability_supplier(
    ws: Worksheet, spec: dict, slot: int, capacity_scores: dict[str, Any]
) -> None:
    s = spec["sheets"]["Capability Evaluation"]
    col = supplier_col_letter(s["supplier_start_col_index"], slot, width_per_slot=1)

    ws[f"{col}{s['supplier_label_row']}"] = f"Soumissionnaire {slot:02d}"
    summary_col = supplier_col_letter(
        spec["sheets"]["Summary"]["supplier_start_col_index"], slot, 1
    )
    ws[f"{col}{s['supplier_name_row']}"] = s["name_formula_pattern"].format(
        summary_col=summary_col
    )

    # Sous-critères : remplir si dispo, sinon laisser vide (human_only styling)
    keys = list(capacity_scores.keys())
    for idx, r in enumerate(range(s["criteria_start_row"], s["criteria_end_row"] + 1)):
        cell = ws[f"{col}{r}"]
        if idx < len(keys):
            val = capacity_scores[keys[idx]]
            if val is None:
                apply_confidence_styling(cell, "human_only")
            else:
                cell.value = val
        else:
            apply_confidence_styling(cell, "human_only")

    ws[f"{col}{s['score_row']}"] = s["score_formula"].format(col=col)


def populate_sustainability_supplier(
    ws: Worksheet, spec: dict, slot: int, sustainability_scores: dict[str, Any]
) -> None:
    s = spec["sheets"]["Sustainability Evaluation"]
    col = supplier_col_letter(s["supplier_start_col_index"], slot, width_per_slot=1)

    ws[f"{col}{s['supplier_label_row']}"] = f"Soumissionnaire {slot:02d}"
    summary_col = supplier_col_letter(
        spec["sheets"]["Summary"]["supplier_start_col_index"], slot, 1
    )
    ws[f"{col}{s['supplier_name_row']}"] = s["name_formula_pattern"].format(
        summary_col=summary_col
    )

    vals = list(sustainability_scores.values())
    for idx, r in enumerate(
        range(s["criteria_start_row"], s["criteria_start_row"] + s["criteria_count"])
    ):
        cell = ws[f"{col}{r}"]
        if idx < len(vals) and vals[idx] is not None:
            cell.value = vals[idx]
        else:
            cell.value = 0
            apply_confidence_styling(cell, "medium")

    ws[f"{col}{s['intermediate_score_row']}"] = s["intermediate_formula"].format(
        col=col
    )
    ws[f"{col}{s['score_row']}"] = s["score_formula"].format(col=col)


def populate_commercial_supplier(
    ws: Worksheet, spec: dict, slot: int, line_items: list[dict[str, Any]]
) -> None:
    s = spec["sheets"]["Commercial Evaluation"]
    price_col, total_col = commercial_cols(s["supplier_start_col_index"], slot)

    summary_col = supplier_col_letter(
        spec["sheets"]["Summary"]["supplier_start_col_index"], slot, 1
    )

    # label+name placed in price_col (first col of block)
    ws[f"{price_col}{s['supplier_label_row']}"] = (
        f"=Summary!{summary_col}{spec['sheets']['Summary']['supplier_label_row']}"
    )
    ws[f"{price_col}{s['supplier_name_row']}"] = (
        f"=Summary!{summary_col}{spec['sheets']['Summary']['supplier_name_row']}"
    )

    # Items: write unit_price and compute totals (leave quantities in column E to be already in template)
    start = s["items_start_row"]
    end = s["items_end_row"]
    for idx, r in enumerate(range(start, end + 1)):
        cell_price = ws[f"{price_col}{r}"]
        cell_total = ws[f"{total_col}{r}"]
        if idx < len(line_items) and line_items[idx].get("unit_price") is not None:
            cell_price.value = line_items[idx]["unit_price"]
        else:
            # leave blank; mark medium (to validate) only if row is used (optional)
            apply_confidence_styling(cell_price, "medium")
        cell_total.value = s["montant_formula"].format(price_col=price_col, row=r)

    ws[f"{total_col}{s['subtotal_row']}"] = s["subtotal_formula"].format(
        total_col=total_col, start=start, end=end
    )
    # Score row left to template formulas, but ensure cell exists
    if ws[f"{total_col}{s['score_row']}"].value is None:
        apply_confidence_styling(ws[f"{total_col}{s['score_row']}"], "medium")
