from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.workbook import Workbook

from .column_calculator import commercial_cols, supplier_col_index
from .supplier_mapper import (
    populate_capability_supplier,
    populate_commercial_supplier,
    populate_essential_supplier,
    populate_summary_supplier,
    populate_sustainability_supplier,
)


@dataclass
class EnginePaths:
    spec_path: Path
    template_path: Path

class TemplateMappingEngine:
    def __init__(self, spec_path: str | Path, template_path: str | Path):
        self.paths = EnginePaths(Path(spec_path), Path(template_path))
        self.spec = self._load_spec(self.paths.spec_path)

    @staticmethod
    def _load_spec(path: Path) -> dict:
        return json.loads(path.read_text(encoding="utf-8"))

    def _load_template(self) -> Workbook:
        if not self.paths.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.paths.template_path}")
        return openpyxl.load_workbook(self.paths.template_path)

    def _validate_num_suppliers(self, n: int) -> int:
        max_s = int(self.spec["max_suppliers"])
        if n < 1:
            raise ValueError("At least 1 submission is required to export a CBA.")
        if n > max_s:
            raise ValueError(f"Too many suppliers: {n}. Max allowed: {max_s}.")
        # UX: show at least min_suppliers_ui slots (no blocking)
        return max(n, int(self.spec.get("min_suppliers_ui", 3)))

    def unmask_suppliers(self, wb: Workbook, n_visible: int) -> None:
        rules = self.spec["expansion"]["rules"]
        max_s = int(self.spec["max_suppliers"])

        for sheet_name, rule in rules.items():
            ws = wb[sheet_name]
            start = int(rule["start_col_index"])
            width = int(rule["width_per_slot"])
            for slot in range(1, max_s + 1):
                # For Commercial, width_per_slot = 2 => needs two columns per slot
                if width == 1:
                    col_idx = supplier_col_index(start, slot, 1)
                    from openpyxl.utils import get_column_letter
                    col = get_column_letter(col_idx)
                    ws.column_dimensions[col].hidden = not (slot <= n_visible)
                else:
                    price_col, total_col = commercial_cols(start, slot)
                    ws.column_dimensions[price_col].hidden = not (slot <= n_visible)
                    ws.column_dimensions[total_col].hidden = not (slot <= n_visible)

    def populate_case(self, wb: Workbook, case_data: Dict[str, Any]) -> Workbook:
        submissions: List[Dict[str, Any]] = case_data.get("submissions", [])
        n = len(submissions)
        n_visible = self._validate_num_suppliers(n)

        # Unmask relevant columns
        self.unmask_suppliers(wb, n_visible)

        # Fill Summary first (names drive formulas)
        ws_summary = wb[self.spec["sheets"]["Summary"]["sheet_name"]]
        for slot, sub in enumerate(submissions, start=1):
            populate_summary_supplier(ws_summary, self.spec, slot, sub.get("supplier_name", f"Soumissionnaire {slot:02d}"))

        # Essential
        ws_ess = wb[self.spec["sheets"]["Essential Evaluation"]["sheet_name"]]
        for slot, sub in enumerate(submissions, start=1):
            populate_essential_supplier(ws_ess, self.spec, slot, sub.get("conformity", {}))

        # Capability
        ws_cap = wb[self.spec["sheets"]["Capability Evaluation"]["sheet_name"]]
        for slot, sub in enumerate(submissions, start=1):
            populate_capability_supplier(ws_cap, self.spec, slot, sub.get("capacity_scores", {}))

        # Sustainability
        ws_sus = wb[self.spec["sheets"]["Sustainability Evaluation"]["sheet_name"]]
        for slot, sub in enumerate(submissions, start=1):
            populate_sustainability_supplier(ws_sus, self.spec, slot, sub.get("sustainability_scores", {}))

        # Commercial
        ws_com = wb[self.spec["sheets"]["Commercial Evaluation"]["sheet_name"]]
        for slot, sub in enumerate(submissions, start=1):
            populate_commercial_supplier(ws_com, self.spec, slot, sub.get("line_items", []))

        return wb

    def export_cba(self, case_data: Dict[str, Any], output_dir: str | Path) -> Path:
        wb = self._load_template()
        wb = self.populate_case(wb, case_data)

        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        case_id = case_data.get("case_id", "CASE")
        version = int(case_data.get("version", 1))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pattern = self.spec["ux"]["export_filename_pattern"]
        filename = pattern.format(case_id=case_id, version=version, timestamp=ts)
        out_path = out_dir / filename

        wb.save(out_path)
        return out_path
