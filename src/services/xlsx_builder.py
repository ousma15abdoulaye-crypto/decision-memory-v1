"""XLSX builder for enterprise PV exports."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill

from src.services.comparative_table_model import (
    build_comparative_table_model_from_snapshot,
)


def _criterion_weight(criteria: list[dict[str, Any]], criterion_id: str) -> float:
    for c in criteria:
        if str(c.get("id")) == str(criterion_id):
            try:
                return float(c.get("weight") or 0)
            except (TypeError, ValueError):
                return 0.0
    return 0.0


def build_xlsx_export(
    snapshot: dict[str, Any], session_id: str, seal_hash: str, sealed_at: str | None
) -> bytes:
    ct = build_comparative_table_model_from_snapshot(snapshot)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparatif"
    ws.freeze_panes = "A2"

    headers = [
        "Fournisseur",
        "Critère",
        "Valeur",
        "Score brut",
        "Poids",
        "Score pondéré export",
        "Confiance",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["F1"].comment = Comment(
        "Colonne « Score pondéré export » : calcul export-only "
        "(non persisté en DB/snapshot).",
        "DMS",
    )

    criteria = ct.get("criteria") or []
    scores = ct.get("scores_matrix") or {}
    bundles = ct.get("bundles") or []

    for bundle in bundles:
        supplier = bundle.get("supplier_name_display") or bundle.get(
            "supplier_name_raw"
        )
        bid = str(bundle.get("id"))
        per_bundle = scores.get(bid, {})
        if not isinstance(per_bundle, dict):
            continue
        for criterion_id, score_payload in per_bundle.items():
            if isinstance(score_payload, dict):
                raw_score = score_payload.get("score") or score_payload.get("value")
                value = score_payload.get("value")
            else:
                raw_score = score_payload
                value = score_payload
            try:
                raw_score_f = float(raw_score)
            except (TypeError, ValueError):
                raw_score_f = 0.0
            weight = _criterion_weight(criteria, str(criterion_id))
            weighted = round(raw_score_f * weight, 4)
            ws.append(
                [
                    supplier,
                    str(criterion_id),
                    value,
                    raw_score_f,
                    weight,
                    weighted,
                    bundle.get("assembly_confidence"),
                ]
            )

    ws.conditional_formatting.add(
        f"F2:F{max(ws.max_row, 2)}",
        DataBarRule(start_type="num", start_value=0, end_type="max", end_value=1),
    )

    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 24
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 16

    trace = wb.create_sheet("Traceability")
    tr = ct.get("trace") or {}
    trace["A1"] = "workspace_id"
    trace["B1"] = snapshot.get("process", {}).get("workspace_id")
    trace["A2"] = "session_id"
    trace["B2"] = session_id
    trace["A3"] = "sealed_at"
    trace["B3"] = sealed_at
    trace["A4"] = "seal_hash_sha256"
    trace["B4"] = seal_hash
    trace["A5"] = "dms_version"
    trace["B5"] = snapshot.get("dms_version")
    trace["A6"] = "snapshot_schema_version"
    trace["B6"] = tr.get("snapshot_schema_version")
    trace["A7"] = "render_template_version"
    trace["B7"] = tr.get("render_template_version")
    trace["A8"] = "comparative_source"
    trace["B8"] = ct.get("source")
    trace.column_dimensions["A"].width = 26
    trace.column_dimensions["B"].width = 90

    m16 = ct.get("m16") if isinstance(ct.get("m16"), dict) else {}
    assessments = m16.get("assessments") or []
    if assessments:
        m16s = wb.create_sheet("M16_assessments")
        m16s.append(["Bundle", "Critère", "Statut", "Confiance", "Cell JSON (aperçu)"])
        for cell in m16s[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for a in assessments:
            cj = a.get("cell_json") if isinstance(a.get("cell_json"), dict) else {}
            preview = str(cj)[:500]
            conf = a.get("confidence")
            m16s.append(
                [
                    str(a.get("bundle_id")),
                    str(a.get("criterion_key")),
                    str(a.get("assessment_status")),
                    conf,
                    preview,
                ]
            )
        m16s.column_dimensions["A"].width = 36
        m16s.column_dimensions["B"].width = 24
        m16s.column_dimensions["C"].width = 18
        m16s.column_dimensions["D"].width = 12
        m16s.column_dimensions["E"].width = 60

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
