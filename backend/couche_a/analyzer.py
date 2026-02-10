"""Couche A – Document extraction engine (PDF, DOCX, XLSX)."""

from __future__ import annotations

import re

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.system.settings import get_settings

__all__ = [
    "extract_from_pdf",
    "extract_from_docx",
    "extract_from_xlsx",
    "extract_vendor_name",
    "extract_amount",
    "detect_doc_type",
    "run_extraction",
]


# ---- Extractors -------------------------------------------------------------

def extract_from_pdf(file_path: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def extract_from_docx(file_path: str) -> str:
    from docx import Document

    doc = Document(file_path)
    return "\n".join(p.text for p in doc.paragraphs)


def extract_from_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            parts.append(" ".join(str(c) for c in row if c is not None))
    wb.close()
    return "\n".join(parts)


# ---- Heuristic extractors ---------------------------------------------------

_VENDOR_RE = re.compile(
    r"(?:soci[ée]t[ée]|entreprise|company|vendor|fournisseur)\s*[:\-]?\s*(.+)",
    re.IGNORECASE,
)


def extract_vendor_name(text: str) -> str | None:
    m = _VENDOR_RE.search(text)
    return m.group(1).strip()[:255] if m else None


_AMOUNT_RE = re.compile(
    r"(?:montant|amount|total|prix)\s*[:\-]?\s*([\d\s.,]+)",
    re.IGNORECASE,
)


def extract_amount(text: str) -> str | None:
    m = _AMOUNT_RE.search(text)
    return m.group(1).strip() if m else None


_DOC_KEYWORDS: dict[str, list[str]] = {
    "DAO": ["dossier d'appel", "cahier des charges", "termes de référence"],
    "TECH": ["offre technique", "mémoire technique", "note méthodologique"],
    "FIN": ["offre financière", "bordereau des prix", "devis", "montant"],
    "AUTRE": [],
}


def detect_doc_type(text: str, filename: str) -> str:
    lower = (text + " " + filename).lower()
    for dtype, keywords in _DOC_KEYWORDS.items():
        if dtype == "AUTRE":
            continue
        if any(kw in lower for kw in keywords):
            return dtype
    return "AUTRE"


# ---- Orchestrator -----------------------------------------------------------

async def run_extraction(document_id: str, file_path: str, db: AsyncSession) -> None:
    """Run extraction on a single document and persist a PreanalysisResult."""
    from backend.couche_a.models import PreanalysisResult, SubmissionDocument

    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
    extractors = {"pdf": extract_from_pdf, "docx": extract_from_docx, "xlsx": extract_from_xlsx}
    extractor = extractors.get(ext)
    if not extractor:
        return

    text = extractor(file_path)
    vendor = extract_vendor_name(text)
    amount = extract_amount(text)
    doc_type = detect_doc_type(text, file_path)
    settings = get_settings()
    llm_used = False

    if settings.USE_LLM:
        # Stub for LLM-enhanced extraction
        llm_used = True

    # Get parent submission id
    doc_stmt = select(SubmissionDocument).where(SubmissionDocument.id == document_id)
    doc = (await db.execute(doc_stmt)).scalar_one_or_none()
    submission_id = doc.submission_id if doc else None

    result = PreanalysisResult(
        submission_id=submission_id or "",
        document_id=document_id,
        vendor_name=vendor,
        amount=amount,
        detected_type=doc_type,
        doc_checklist={"has_technical": doc_type == "TECH", "has_financial": doc_type == "FIN"},
        flags={},
        llm_used=llm_used,
    )
    db.add(result)
    await db.flush()
