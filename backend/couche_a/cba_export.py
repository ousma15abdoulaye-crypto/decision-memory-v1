"""Couche A – CBA (Comparative Bid Analysis) Excel export and import."""

from __future__ import annotations

import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.couche_a.models import PreanalysisResult, Submission
from backend.couche_a.rules_engine import evaluate_submission
from backend.system.settings import get_settings

__all__ = ["generate_cba_excel", "parse_revised_cba"]

ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
HEADER_FONT = Font(bold=True)


async def generate_cba_excel(case_id: str, lot_id: str, db: AsyncSession) -> str:
    """Generate a multi-tab CBA workbook and return the output file path."""
    settings = get_settings()
    output_dir = os.path.join(settings.OUTPUT_DIR, case_id, lot_id)
    os.makedirs(output_dir, exist_ok=True)

    stmt = select(Submission).where(
        Submission.case_id == case_id, Submission.lot_id == lot_id
    )
    submissions = (await db.execute(stmt)).scalars().all()

    evaluations: list[dict] = []
    for sub in submissions:
        ev = await evaluate_submission(sub.id, db)
        evaluations.append(ev)

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = ["Vendor", "Status", "Total", "Capacity", "Durability", "Commercial"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT

    for row_idx, ev in enumerate(evaluations, 2):
        sub_stmt = select(Submission).where(Submission.id == ev["submission_id"])
        sub = (await db.execute(sub_stmt)).scalar_one_or_none()
        vendor = sub.vendor_name if sub else "Unknown"
        ws_summary.cell(row=row_idx, column=1, value=vendor)
        status_cell = ws_summary.cell(row=row_idx, column=2, value=ev["status"])
        ws_summary.cell(row=row_idx, column=3, value=ev["total"])
        ws_summary.cell(row=row_idx, column=4, value=ev["scores"]["capacity"])
        ws_summary.cell(row=row_idx, column=5, value=ev["scores"]["durability"])
        ws_summary.cell(row=row_idx, column=6, value=ev["scores"]["commercial"])
        if ev["status"] == "REVUE_MANUELLE":
            status_cell.fill = ORANGE_FILL

    # --- Essential sheet ---
    ws_ess = wb.create_sheet("Essential")
    ws_ess.cell(row=1, column=1, value="Vendor").font = HEADER_FONT
    ws_ess.cell(row=1, column=2, value="Pass").font = HEADER_FONT
    ws_ess.cell(row=1, column=3, value="Reasons").font = HEADER_FONT
    for row_idx, ev in enumerate(evaluations, 2):
        sub_stmt = select(Submission).where(Submission.id == ev["submission_id"])
        sub = (await db.execute(sub_stmt)).scalar_one_or_none()
        ws_ess.cell(row=row_idx, column=1, value=sub.vendor_name if sub else "Unknown")
        ws_ess.cell(row=row_idx, column=2, value="YES" if ev["essential_pass"] else "NO")
        ws_ess.cell(row=row_idx, column=3, value="; ".join(ev["essential_reasons"]))

    # --- Capability sheet ---
    ws_cap = wb.create_sheet("Capability")
    ws_cap.cell(row=1, column=1, value="Vendor").font = HEADER_FONT
    ws_cap.cell(row=1, column=2, value="Capacity Score").font = HEADER_FONT
    for row_idx, ev in enumerate(evaluations, 2):
        sub_stmt = select(Submission).where(Submission.id == ev["submission_id"])
        sub = (await db.execute(sub_stmt)).scalar_one_or_none()
        ws_cap.cell(row=row_idx, column=1, value=sub.vendor_name if sub else "Unknown")
        ws_cap.cell(row=row_idx, column=2, value=ev["scores"]["capacity"])

    # --- Sustainability sheet ---
    ws_dur = wb.create_sheet("Sustainability")
    ws_dur.cell(row=1, column=1, value="Vendor").font = HEADER_FONT
    ws_dur.cell(row=1, column=2, value="Durability Score").font = HEADER_FONT
    for row_idx, ev in enumerate(evaluations, 2):
        sub_stmt = select(Submission).where(Submission.id == ev["submission_id"])
        sub = (await db.execute(sub_stmt)).scalar_one_or_none()
        ws_dur.cell(row=row_idx, column=1, value=sub.vendor_name if sub else "Unknown")
        ws_dur.cell(row=row_idx, column=2, value=ev["scores"]["durability"])

    # --- Commercial sheet ---
    ws_com = wb.create_sheet("Commercial")
    ws_com.cell(row=1, column=1, value="Vendor").font = HEADER_FONT
    ws_com.cell(row=1, column=2, value="Commercial Score").font = HEADER_FONT
    for row_idx, ev in enumerate(evaluations, 2):
        sub_stmt = select(Submission).where(Submission.id == ev["submission_id"])
        sub = (await db.execute(sub_stmt)).scalar_one_or_none()
        ws_com.cell(row=row_idx, column=1, value=sub.vendor_name if sub else "Unknown")
        ws_com.cell(row=row_idx, column=2, value=ev["scores"]["commercial"])

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"CBA_{case_id}_{lot_id}_{ts}.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def parse_revised_cba(file_path: str) -> dict:
    """Parse a committee-revised CBA workbook and return override dict."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    revisions: dict = {}

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if row and len(row) >= 6:
                vendor = row[0]
                revisions[vendor] = {
                    "status": row[1],
                    "total": row[2],
                    "capacity": row[3],
                    "durability": row[4],
                    "commercial": row[5],
                }

    wb.close()
    return revisions
